import socket
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

HOST = '192.168.1.236'  # Standard loopback interface address (localhost)
PORT = 8000        # Port to listen on (non-privileged ports are > 1023)

wb = load_workbook('data.xlsx')
sheet = wb['Sheet1']
i = 1
while True:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((HOST, PORT))
        s.listen()
        conn, addr = s.accept()
        with conn:

            print('Connected by', addr)
            while True:
                s.settimeout(1000)
                data = conn.recv(1024)
                s.settimeout(None)
                if not data:
                    wb.save('data.xlsx')  
                    break
                print('Data = ', data)
                data1 = data
               
                sheet.cell(row=i, column=1, value=data1)
                i = i+1
                conn.sendall(data1)
                
df = pd.read_excel('data.xlsx')
answer = df[1:57]
print(answer)
plt.xlabel("time(s)")
plt.ylabel("degC")
plt.title('TEMPERATURE vs TIME')
x_ticks = np.arange(0,40,2)
y_ticks = np.arange(25,35,0.5)
plt.xticks(x_ticks)
plt.yticks(y_ticks)
plt.plot(range(1,len(answer)+1),answer)
plt.show()      

        
                
                

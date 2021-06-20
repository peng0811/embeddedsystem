from bluepy.btle import Peripheral, UUID
from bluepy.btle import Scanner, DefaultDelegate
from openpyxl import load_workbook
import time
wb=load_workbook
sheet=wb['Sheet1']
idx=2

class ScanDelegate(DefaultDelegate):
	def __init__(self):
		DefaultDelegate.__init__(self)
	def handleDiscovery(self, dev, isNewDev, isNewData):
		if isNewDev:
			print ("Discovered device", dev.addr)
		elif isNewData:
			print ("Received new data from", dev.addr)

scanner = Scanner().withDelegate(ScanDelegate())
devices = scanner.scan(10.0)



n=0
for dev in devices:
	print ("%d: Device %s (%s), RSSI=%d dB" % (n, dev.addr,dev.addrType, dev.rssi))
	n = n+1
	for (adtype, desc, value) in dev.getScanData():
		print (" %s = %s" % (desc, value))

number = input('Enter your device number: ')
print('Device', number)

number = int(number)
n=0
for dev in devices:
	print ("%d: Device %s (%s), RSSI=%d dB" % (n, dev.addr,dev.addrType, dev.rssi))
	if n == number:
		word = dev.addr
	n = n+1
	for (adtype, desc, value) in dev.getScanData():
		print (" %s = %s" % (desc, value))


print(word)
print ("Connecting...")
dev = Peripheral(word, 'random')
print ("Services...")
for svc in dev.services:
	print (str(svc))

try:
	testService = dev.getServiceByUUID(UUID(0xfff0))
	for ch in testService.getCharacteristics():
		print (str(ch))
	for i in range(1,50):
		ch = dev.getCharacteristics(uuid=UUID(0xfff1))[0]
		if (ch.supportsRead()):
		    c = ch.read()
		    if c[0] >= 35:
			print (ch.read())
			time.sleep(2)
			data = c[0]
                        sheet.cell(row=idx,column=2,value=data)
                        idx=idx+1
                        wb.save('data3.xlsx')
                        break

finally:
	dev.disconnect()

from bluepy.btle import Peripheral, UUID
from bluepy.btle import Scanner, DefaultDelegate
from openpyxl import load_workbook
import time
from random import randint
import asyncio
import socket
import json
import random
import websockets
import time
import threading
wb=load_workbook('data3.xlsx')
sheet=wb['Sheet1']
idx=2
class ScanDelegate(DefaultDelegate):
	def __init__(self):
		DefaultDelegate.__init__(self)
	def handleDiscovery(self, dev, isNewDev, isNewData):
		if isNewDev:
			print ("Discovered device", dev.addr)
		elif isNewData:
			print ("Received new data from", dev.addr)

scanner = Scanner().withDelegate(ScanDelegate())
devices = scanner.scan(10.0)
num=0
run = True
async def web_socket(websocket,path):
    global num
    global data
    global times
    global idx
    times = 0
    while(run):
        s = await websocket.recv()
        print(s)
        if s!=1:
            sheet.cell(row=idx,column=1,value=s)
            wb.save('data3.xlsx')
            while 1:
                data_rec()
                if num >=350:
                    times += 1
                print(times)
                data= "{\"val\":%d , \"time\":%d }" % (num,times)
                time.sleep(1)
                await websocket.send(data)
                if num >=350:
                    sheet.cell(row=idx,column=2,value=num/10)
                    wb.save('data3.xlsx')
                    idx=idx+1
                    break
    

n=0
for dev in devices:
	print ("%d: Device %s (%s), RSSI=%d dB" % (n, dev.addr,dev.addrType, dev.rssi))
	n = n+1
	for (adtype, desc, value) in dev.getScanData():
		print (" %s = %s" % (desc, value))

number = input('Enter your device number: ')
print('Device', number)

number = int(number)
n=0
for dev in devices:
	print ("%d: Device %s (%s), RSSI=%d dB" % (n, dev.addr,dev.addrType, dev.rssi))
	if n == number:
		word = dev.addr
	n = n+1
	for (adtype, desc, value) in dev.getScanData():
		print (" %s = %s" % (desc, value))

print(word)
print ("Connecting...")
dev = Peripheral(word, 'random')
print ("Services...")
for svc in dev.services:
    print (str(svc))
k=1
try:
    testService = dev.getServiceByUUID(UUID(0xfff0))
    for ch in testService.getCharacteristics():
        print (str(ch))
        ch = dev.getCharacteristics(uuid=UUID(0xfff1))[0]
        def data_rec():
            global num
            global ID
            if (ch.supportsRead()):
                c = ch.read()
                num =(int(c[0])+256)/10
                num=num*10
                time.sleep(0.5)

    #t1 = threading.Thread(target=data_rec)
    #t1.start()

    start_server = websockets.serve(web_socket,"127.0.0.1",5050)
    asyncio.get_event_loop().run_until_complete(start_server)
    asyncio.get_event_loop().run_forever()
finally:
	dev.disconnect()
